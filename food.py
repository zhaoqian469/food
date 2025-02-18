import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, timedelta
import chinese_calendar as calendar

# 获取指定年份的法定节假日
def get_holidays(year):
    holidays = []
    start_date = datetime(year, 1, 1)
    end_date = datetime(year, 12, 31)
    current_date = start_date

    while current_date <= end_date:
        if calendar.is_holiday(current_date):
            holidays.append(current_date)
        current_date += timedelta(days=1)

    return pd.to_datetime(holidays)

# 定义餐补时间段
def get_meal_period(time):
    if pd.to_datetime("07:20:00").time() <= time <= pd.to_datetime("09:00:00").time():
        return "早餐"
    elif pd.to_datetime("11:00:00").time() <= time <= pd.to_datetime("14:00:00").time():
        return "午餐"
    elif pd.to_datetime("17:00:00").time() <= time <= pd.to_datetime("20:00:00").time():
        return "晚餐"
    else:
        return "其他"

# 计算餐补金额
def calculate_subsidy(group):
    subsidy_used = 0  # 记录已使用的补贴
    max_subsidy = 0  # 记录当前时段最大可用补贴

    for index, row in group.iterrows():
        if row["交易地点"] == "超市":  # 如果交易地点为超市，跳过餐补计算
            df.at[index, "餐补金额"] = 0
            df.at[index, "自付（元）"] = row["交易金额"]
            df.at[index, "早餐（元）"] = 0
            df.at[index, "工作餐（元）"] = 0
            df.at[index, "加班餐（元）"] = 0
            continue  # 跳过超市的餐补计算

        date = row["交易时间"].date()
        weekday = row["交易时间"].weekday()
        meal_period = row["餐费时间段"]

        workday = (weekday < 5) or (date in overtime_dates)
        is_holiday = (date in HOLIDAY_AND_HIGH_TEMP_DAYS) and (date not in overtime_dates)

        # 确定餐补上限
        if row["人员类别"] == "职工":
            if meal_period == "早餐":
                max_subsidy = 0
            elif meal_period in ["午餐", "晚餐"] and is_holiday:
                max_subsidy = 29
            elif meal_period == "午餐" and workday:
                max_subsidy = 25
            elif meal_period in ["午餐", "晚餐"]:
                max_subsidy = 29
            elif meal_period not in ["早餐", "午餐", "晚餐"]:
                max_subsidy = 0
        elif row["人员类别"] == "研究生":
            if meal_period == "早餐" and workday:
                max_subsidy = 2
            elif meal_period in ["午餐", "晚餐"] and is_holiday:
                max_subsidy = 29
            elif meal_period == "午餐" and workday:
                max_subsidy = 25
            elif meal_period in ["午餐", "晚餐"]:
                max_subsidy = 29
            elif meal_period not in ["早餐", "午餐", "晚餐"]:
                max_subsidy = 0

        # 计算当前交易可用餐补
        available_subsidy = max(0, max_subsidy - subsidy_used)
        if row["交易金额"] > available_subsidy:
            subsidy_given = available_subsidy
        else:
            subsidy_given = row["交易金额"]

        # 更新补贴已使用金额
        subsidy_used += subsidy_given

        # 计算餐补金额和自付金额
        df.at[index, "餐补金额"] = subsidy_given
        df.at[index, "自付（元）"] = row["交易金额"] - subsidy_given

        # 根据就餐时段分类餐补金额
        if meal_period == "早餐":
            df.at[index, "早餐（元）"] = subsidy_given
        elif meal_period == "午餐" and workday:
            df.at[index, "工作餐（元）"] = subsidy_given
        elif meal_period in ["午餐", "晚餐"]:
            df.at[index, "加班餐（元）"] = subsidy_given

# Streamlit 页面
st.title("餐补计算小程序")

# 上传 CSV 文件
uploaded_file = st.file_uploader("上传 CSV 文件", type=["csv"])
if uploaded_file is not None:
    # 读取数据
    df = pd.read_csv(uploaded_file, encoding="gbk")

    # 用户输入节假日年份
    holiday_year = st.number_input("请输入节假日年份", min_value=2000, max_value=2100, value=2024, step=1)

    # 获取节假日
    holidays = get_holidays(holiday_year)

    # 用户输入加班调休日期
    overtime_dates_input = st.text_input("请输入加班调休日期（格式：YYYY-MM-DD,YYYY-MM-DD,...）", value="2024-03-15,2024-03-16")
    overtime_dates = {datetime.strptime(date.strip(), "%Y-%m-%d").date() for date in overtime_dates_input.split(",")}

    # 用户输入高温假日期
    high_temp_days_input = st.text_input("请输入高温假日期（格式：YYYY-MM-DD,YYYY-MM-DD,...）", value="2024-07-15,2024-07-16")
    high_temp_days = {datetime.strptime(date.strip(), "%Y-%m-%d").date() for date in high_temp_days_input.split(",") if
                      date.strip()}

    # 合并法定节假日和高温假
    HOLIDAY_AND_HIGH_TEMP_DAYS = set(holidays.date).union(high_temp_days)

    # 选择所需字段
    columns_needed = ["人员类别", "姓名", "个人编号", "卡片类型", "交易地点", "交易金额", "交易时间", "卡户部门", "交易类型"]
    df = df[columns_needed]

    # 删除 "收费冲正" 交易记录
    df = df[df["交易类型"] != "收费冲正"]

    # 交易金额转正数
    df["交易金额"] = df["交易金额"].abs()
    df["交易时间"] = pd.to_datetime(df["交易时间"], errors="coerce")

    # 判断职工还是研究生
    df["餐费时间段"] = df["交易时间"].apply(lambda x: get_meal_period(x.time()))

    # 初始化新列
    df["餐补金额"] = 0
    df["自付（元）"] = 0
    df["早餐（元）"] = 0
    df["工作餐（元）"] = 0
    df["加班餐（元）"] = 0

    # 按日期、姓名、个人编号、餐费时间段分组计算餐补
    df = df.sort_values(by=["姓名", "个人编号", "交易时间"])
    df.groupby(["姓名", "个人编号", "交易时间", "餐费时间段"]).apply(calculate_subsidy)

    # 选择最终字段
    df_final = df[["人员类别", "姓名", "个人编号", "卡片类型", "交易地点", "卡户部门", "交易时间", "交易金额",
                   "早餐（元）", "工作餐（元）", "加班餐（元）", "自付（元）"]]

    # 保存结果为 Excel 文件并处理列宽和筛选
    output_file = "./餐补计算结果.xlsx"
    df_final.to_excel(output_file, index=False)

    # === 在 Excel 中自动调整列宽，并添加筛选 ===
    wb = load_workbook(output_file)
    ws = wb.active

    # 设置自动筛选，仅针对 "交易地点" 和 "卡户部门"
    ws.auto_filter.ref = "E1:F" + str(ws.max_row)  # E: 交易地点, F: 卡户部门

    # 设置列宽
    for col in ws.columns:
        col_letter = col[0].column_letter

        # 时间列宽度固定
        if col_letter == "E" or col_letter == "G":  # 交易时间列
            ws.column_dimensions[col_letter].width = 20
        elif col_letter == "F":
            ws.column_dimensions[col_letter].width = 27
        else:
            ws.column_dimensions[col_letter].width = 13  # 最小宽度10

    # 保存调整后的 Excel 文件
    wb.save(output_file)

    # 显示结果
    st.write("✅ 数据处理完成！")
    st.dataframe(df_final)

    # 下载按钮
    st.download_button("📥 下载 Excel 文件", data=open(output_file, "rb").read(), file_name="餐补计算结果.xlsx")
