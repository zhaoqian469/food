import streamlit as st
import pandas as pd
import chinese_calendar as calendar
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook


# 获取指定年份的法定节假日
def get_holidays(year):
    holidays = []
    start_date = datetime(year, 1, 1)
    end_date = datetime(year, 12, 31)
    current_date = start_date

    while current_date <= end_date:
        if calendar.is_holiday(current_date):
            holidays.append(current_date)
        current_date += timedelta(days=1)

    return pd.to_datetime(holidays)


# 定义时间段标签
def get_meal_period(time):
    if pd.to_datetime("07:20:00").time() <= time <= pd.to_datetime("09:00:00").time():
        return "早餐"
    elif pd.to_datetime("11:00:00").time() <= time <= pd.to_datetime("14:00:00").time():
        return "午餐"
    elif pd.to_datetime("17:00:00").time() <= time <= pd.to_datetime("20:00:00").time():
        return "晚餐"
    else:
        return "其他"


# 设定补贴规则
def classify_meal(row, total_amount):
    date = row["交易时间"].date()
    weekday = row["交易时间"].weekday()
    meal_period = row["餐费时间段"]
    is_holiday = date in HOLIDAY_AND_HIGH_TEMP_DAYS or weekday >= 5
    subsidy_limit = 0

    if row["身份"] == "职工":
        if meal_period == "午餐" and weekday < 5:
            subsidy_limit = 25
        elif meal_period in ["午餐", "晚餐"] and is_holiday:
            subsidy_limit = 29
        elif meal_period == "晚餐":
            subsidy_limit = 29
    elif row["身份"] == "学生":
        if meal_period == "早餐" and weekday < 5:
            subsidy_limit = 2
        elif meal_period == "午餐" and weekday < 5:
            subsidy_limit = 25
        elif meal_period in ["午餐", "晚餐"] and is_holiday:
            subsidy_limit = 29
        elif meal_period == "晚餐":
            subsidy_limit = 29

    extra_payment = max(0, total_amount - subsidy_limit) if row["是否最后一笔"] else 0
    return subsidy_limit, extra_payment


# Streamlit 页面
st.title("餐补计算小程序")

# 上传 CSV 文件
uploaded_file = st.file_uploader("上传 CSV 文件", type=["csv"])
if uploaded_file is not None:
    # 读取数据
    df = pd.read_csv(uploaded_file, encoding="gbk")

    # 设定高温假时间范围
    start_date = st.date_input("选择高温假开始日期", value=datetime(2024, 7, 27))
    end_date = st.date_input("选择高温假结束日期", value=datetime(2024, 8, 4))
    HIGH_TEMP_DAYS = pd.date_range(start=start_date, end=end_date)

    # 获取节假日
    HOLIDAYS_2024 = get_holidays(2024)
    HOLIDAYS_2025 = get_holidays(2025)
    HOLIDAY_AND_HIGH_TEMP_DAYS_2024 = pd.to_datetime(HIGH_TEMP_DAYS).union(pd.to_datetime(HOLIDAYS_2024))
    HOLIDAY_AND_HIGH_TEMP_DAYS = pd.to_datetime(HOLIDAY_AND_HIGH_TEMP_DAYS_2024).union(pd.to_datetime(HOLIDAYS_2025))

    # 选择所需字段
    columns_needed = ["帐号", "姓名", "个人编号", "卡片类型", "交易地点", "交易金额", "交易时间", "卡户部门", "交易类型"]
    df = df[columns_needed]

    # **✅ 删除 "收费冲正" 交易记录**
    df = df[df["交易类型"] != "收费冲正"]

    # 交易金额转正数
    df["交易金额"] = df["交易金额"].abs()
    df["交易时间"] = pd.to_datetime(df["交易时间"], errors="coerce")

    # 判断职工还是学生
    df["身份"] = df["帐号"].astype(str).apply(lambda x: "职工" if len(x) == 4 else "学生" if len(x) == 8 else "未知")
    df["餐费时间段"] = df["交易时间"].apply(lambda x: get_meal_period(x.time()))
    df["日期"] = df["交易时间"].dt.date
    df["是否最后一笔"] = df.duplicated(subset=["姓名", "个人编号", "日期", "餐费时间段"], keep="last") == False
    df["总交易金额"] = df.groupby(["姓名", "个人编号", "日期", "餐费时间段"])["交易金额"].transform("sum")

    # 计算补贴和超额
    df[["补贴上限", "自付（元）"]] = df.apply(lambda row: pd.Series(classify_meal(row, row["总交易金额"])), axis=1)

    # 计算餐费类别
    df["工作餐（元）"] = df.apply(lambda x: x["交易金额"] if x["餐费时间段"] == "午餐" and x["补贴上限"] > 0 else 0, axis=1)
    df["加班餐（元）"] = df.apply(lambda x: x["交易金额"] if x["餐费时间段"] == "晚餐" and x["补贴上限"] > 0 else 0, axis=1)
    df["早餐（元）"] = df.apply(lambda x: x["交易金额"] if x["餐费时间段"] == "早餐" and x["补贴上限"] > 0 else 0, axis=1)

    # 选择最终字段
    df_final = df[["姓名", "个人编号", "卡片类型", "交易地点", "卡户部门", "交易时间", "交易金额",
                   "早餐（元）", "工作餐（元）", "加班餐（元）", "自付（元）"]]

    # 显示结果
    st.write("✅ 数据处理完成！")
    st.dataframe(df_final)

    # 保存 Excel 文件并提供下载
    output_file = BytesIO()
    df_final.to_excel(output_file, index=False, engine="openpyxl")
    output_file.seek(0)

    # === 在 Excel 中自动调整列宽，并添加筛选 ===
    wb = load_workbook(output_file)
    ws = wb.active

    # 设置自动筛选
    ws.auto_filter.ref = "D1:E1" + str(ws.max_row)

    # 设置列宽
    for col in ws.columns:
        col_letter = col[0].column_letter

        # 时间列宽度固定
        if col_letter == "D" or col_letter == "F":  # 交易时间列
            ws.column_dimensions[col_letter].width = 20
        elif col_letter == "E":
            ws.column_dimensions[col_letter].width = 27
        else:
            ws.column_dimensions[col_letter].width = 10  # 最小宽度10

    # 保存 Excel 文件
    output_file.seek(0)
    wb.save(output_file)
    output_file.seek(0)

    # 下载按钮
    st.download_button(
        label="📥 下载 Excel 文件",
        data=output_file,
        file_name="结果.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
